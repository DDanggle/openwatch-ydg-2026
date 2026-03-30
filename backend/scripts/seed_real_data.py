#!/usr/bin/env python3
"""2026 재산공개 엑셀 → SQLite 실데이터 시딩 + 실거래가 매칭."""

import asyncio
import hashlib
import json
import logging
import math
import sqlite3
import statistics
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.database import get_db, init_db
from app.ingestion.address_parser import find_lawd_code, match_trade_record, parse_address
from app.ingestion.real_estate import fetch_trades_for_period

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger(__name__)

EXCEL_PATH = Path(__file__).resolve().parent.parent.parent / "data" / "2026 국회고위공직자 재산정보(공개본).xlsx"

# 금액 변환: 엑셀은 천원, DB는 원
UNIT_MULTIPLIER = 1_000


def _make_id(name: str, position: str) -> str:
    """성명+직위로 고유 ID 생성."""
    raw = f"{name}_{position}"
    return "P" + hashlib.md5(raw.encode()).hexdigest()[:7]


def seed_politicians(conn: sqlite3.Connection) -> dict[str, str]:
    """엑셀 총계 시트에서 의원 목록 추출."""
    wb = openpyxl.load_workbook(str(EXCEL_PATH), read_only=True)
    ws = wb["총계"]

    name_to_id: dict[str, str] = {}
    count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        cols = list(row)
        if cols[0] is None:
            continue

        party = str(cols[1] or "")
        position = str(cols[2] or "")
        name = str(cols[3] or "").strip()

        if not name:
            continue

        pid = _make_id(name, position)
        name_to_id[name] = pid

        # 총재산 정보
        total_prev = int((cols[4] or 0) * UNIT_MULTIPLIER)
        total_curr = int((cols[7] or 0) * UNIT_MULTIPLIER)

        conn.execute(
            """INSERT OR REPLACE INTO politician_master
               (id, name, party, position, raw_json)
               VALUES (?, ?, ?, ?, ?)""",
            (
                pid, name, party, position,
                json.dumps({
                    "총_종전가액": total_prev,
                    "총_현재가액": total_curr,
                }, ensure_ascii=False),
            ),
        )
        count += 1

    wb.close()
    logger.info(f"의원 {count}명 입력 완료")
    return name_to_id


def seed_assets(conn: sqlite3.Connection, name_to_id: dict[str, str]) -> int:
    """엑셀 상세 시트에서 자산 항목 추출."""
    wb = openpyxl.load_workbook(str(EXCEL_PATH), read_only=True)
    ws = wb["상세"]

    count = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        cols = list(row)
        if cols[0] is None:
            continue

        name = str(cols[3] or "").strip()
        pid = name_to_id.get(name)
        if not pid:
            skipped += 1
            continue

        asset_type = str(cols[4] or "")
        relation = str(cols[5] or "")
        kind = str(cols[6] or "")
        detail = str(cols[7] or "")

        def safe_int(val) -> int:
            if val is None:
                return 0
            if isinstance(val, (int, float)):
                return int(val * UNIT_MULTIPLIER)
            return 0

        origin = safe_int(cols[8])
        increased = safe_int(cols[9])
        decreased = safe_int(cols[11])
        current = safe_int(cols[13])
        reason = str(cols[14] or "") if len(cols) > 14 else ""

        # 숫자가 아닌 행 (구분선 등) 스킵
        if current == 0 and origin == 0 and not detail:
            skipped += 1
            continue

        inc_real = safe_int(cols[10])
        dec_real = safe_int(cols[12])

        raw = json.dumps({
            "증가액실거래가격": inc_real,
            "감소액실거래가격": dec_real,
        }, ensure_ascii=False)

        try:
            conn.execute(
                """INSERT OR REPLACE INTO asset_itemized
                   (politician_id, disclosure_date, disclosure_year,
                    asset_type, relation, kind, detail,
                    origin_valuation, increased_amount, decreased_amount,
                    current_valuation, reason, raw_json)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    pid, "202603", 2026,
                    asset_type, relation, kind, detail,
                    origin, increased, decreased, current,
                    reason, raw,
                ),
            )
            count += 1
        except sqlite3.IntegrityError:
            # 중복 키 (같은 의원, 같은 자산 항목)
            skipped += 1

    wb.close()
    logger.info(f"자산 항목 {count}건 입력 (스킵: {skipped})")
    return count


def create_synthetic_prior_year(conn: sqlite3.Connection):
    """종전가액으로 합성 2025년 데이터 생성 → YoY 계산 가능하게."""
    conn.execute(
        """INSERT OR IGNORE INTO asset_itemized
           (politician_id, disclosure_date, disclosure_year,
            asset_type, relation, kind, detail,
            origin_valuation, increased_amount, decreased_amount,
            current_valuation, reason, raw_json)
           SELECT
            politician_id, '202503', 2025,
            asset_type, relation, kind, detail,
            0, 0, 0,
            origin_valuation,  -- 종전가액을 전년도 현재가액으로
            '합성(종전가액 기반)', '{}'
           FROM asset_itemized
           WHERE disclosure_year = 2026
             AND origin_valuation > 0"""
    )
    cnt = conn.execute(
        "SELECT COUNT(*) as c FROM asset_itemized WHERE disclosure_year = 2025"
    ).fetchone()["c"]
    logger.info(f"합성 2025년 데이터 {cnt}건 생성")


def compute_yearly_aggregates(conn: sqlite3.Connection) -> int:
    """연도별 집계."""
    conn.execute("DELETE FROM asset_disclosure_yearly")
    conn.execute("""
        INSERT INTO asset_disclosure_yearly (
            politician_id, disclosure_year,
            total_assets, total_debt, net_assets,
            real_estate_total, deposit_total, securities_total,
            self_total, spouse_total, family_total
        )
        SELECT
            politician_id, disclosure_year,
            SUM(CASE WHEN asset_type NOT IN ('채무','채권') THEN current_valuation ELSE 0 END),
            SUM(CASE WHEN asset_type IN ('채무') THEN current_valuation ELSE 0 END),
            SUM(CASE WHEN asset_type NOT IN ('채무','채권') THEN current_valuation ELSE 0 END)
              - SUM(CASE WHEN asset_type IN ('채무') THEN current_valuation ELSE 0 END),
            SUM(CASE WHEN asset_type IN ('건물','토지') THEN current_valuation ELSE 0 END),
            SUM(CASE WHEN asset_type = '예금' THEN current_valuation ELSE 0 END),
            SUM(CASE WHEN asset_type = '증권' THEN current_valuation ELSE 0 END),
            SUM(CASE WHEN relation = '본인' AND asset_type NOT IN ('채무','채권')
                THEN current_valuation ELSE 0 END),
            SUM(CASE WHEN relation = '배우자'
                THEN current_valuation ELSE 0 END),
            SUM(CASE WHEN relation != '본인'
                THEN current_valuation ELSE 0 END)
        FROM asset_itemized
        GROUP BY politician_id, disclosure_year
    """)

    # YoY
    conn.execute("""
        UPDATE asset_disclosure_yearly AS cur
        SET yoy_change = cur.net_assets - (
                SELECT prev.net_assets FROM asset_disclosure_yearly prev
                WHERE prev.politician_id = cur.politician_id
                  AND prev.disclosure_year = cur.disclosure_year - 1
            ),
            yoy_change_pct = CASE
                WHEN (SELECT prev.net_assets FROM asset_disclosure_yearly prev
                      WHERE prev.politician_id = cur.politician_id
                        AND prev.disclosure_year = cur.disclosure_year - 1) != 0
                THEN (cur.net_assets - (
                    SELECT prev.net_assets FROM asset_disclosure_yearly prev
                    WHERE prev.politician_id = cur.politician_id
                      AND prev.disclosure_year = cur.disclosure_year - 1
                )) * 100.0 / ABS((
                    SELECT prev.net_assets FROM asset_disclosure_yearly prev
                    WHERE prev.politician_id = cur.politician_id
                      AND prev.disclosure_year = cur.disclosure_year - 1
                ))
                ELSE 0
            END
        WHERE EXISTS (
            SELECT 1 FROM asset_disclosure_yearly prev
            WHERE prev.politician_id = cur.politician_id
              AND prev.disclosure_year = cur.disclosure_year - 1
        )
    """)

    cnt = conn.execute("SELECT COUNT(*) as c FROM asset_disclosure_yearly").fetchone()["c"]
    logger.info(f"연도별 집계 {cnt}건 완료")
    return cnt


async def match_real_estate(conn: sqlite3.Connection) -> int:
    """아파트 항목을 실거래가와 매칭."""
    # 아파트 항목만 가져오기
    items = conn.execute("""
        SELECT id, politician_id, detail, current_valuation, disclosure_year
        FROM asset_itemized
        WHERE disclosure_year = 2026
          AND asset_type = '건물'
          AND kind LIKE '%아파트%'
          AND detail IS NOT NULL AND detail != ''
    """).fetchall()

    logger.info(f"아파트 매칭 대상: {len(items)}건")

    trade_cache: dict[str, list[dict]] = {}
    matched = 0

    for i, item in enumerate(items):
        parsed = parse_address(item["detail"])
        lawd_cd = find_lawd_code(parsed)

        if not lawd_cd:
            continue

        # 캐시에서 거래 데이터 가져오기
        if lawd_cd not in trade_cache:
            try:
                trades = await fetch_trades_for_period(lawd_cd, "202404", "202503")
                trade_cache[lawd_cd] = trades
                logger.info(f"  [{i+1}/{len(items)}] {lawd_cd}: {len(trades)}건 거래 데이터")
            except Exception as e:
                logger.warning(f"  [{i+1}/{len(items)}] {lawd_cd} 실패: {e}")
                trade_cache[lawd_cd] = []

        trades = trade_cache.get(lawd_cd, [])
        match = match_trade_record(parsed, trades)

        if match.confidence > 0 and match.trade_price:
            declared = item["current_valuation"]
            estimated = match.trade_price
            gap = estimated - declared
            gap_pct = (gap / declared * 100) if declared > 0 else 0

            conn.execute(
                """INSERT OR REPLACE INTO real_estate_match
                   (asset_item_id, politician_id,
                    sido, sigungu, dong, apartment_name, area_sqm,
                    matched_trade_price, match_confidence, match_method,
                    declared_valuation, estimated_market_value,
                    gap_amount, gap_pct)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    item["id"], item["politician_id"],
                    parsed.sido_short, parsed.sigungu, parsed.dong,
                    parsed.apartment_name or match.apartment_name,
                    parsed.area_sqm or match.area_sqm,
                    match.trade_price, match.confidence, match.method,
                    declared, estimated, gap, round(gap_pct, 2),
                ),
            )
            matched += 1

    logger.info(f"실거래가 매칭 완료: {matched}건")
    return matched


def compute_derived_metrics(conn: sqlite3.Connection) -> int:
    """파생 지표 계산."""
    conn.execute("DELETE FROM derived_metrics")

    politicians = conn.execute(
        "SELECT DISTINCT politician_id FROM asset_disclosure_yearly"
    ).fetchall()

    all_yoy = [r["yoy_change"] for r in conn.execute(
        "SELECT yoy_change FROM asset_disclosure_yearly WHERE yoy_change != 0 AND yoy_change IS NOT NULL"
    ).fetchall()]

    median_yoy = statistics.median(all_yoy) if all_yoy else 0
    stdev_yoy = statistics.stdev(all_yoy) if len(all_yoy) > 2 else 1

    count = 0
    for row in politicians:
        pid = row["politician_id"]

        yearly = conn.execute("""
            SELECT * FROM asset_disclosure_yearly
            WHERE politician_id = ? ORDER BY disclosure_year
        """, (pid,)).fetchall()

        if not yearly:
            continue

        first, last = yearly[0], yearly[-1]
        years = last["disclosure_year"] - first["disclosure_year"]

        abs_growth = last["net_assets"] - first["net_assets"]
        growth_rate = (abs_growth / abs(first["net_assets"]) * 100) if first["net_assets"] != 0 else 0

        cagr = None
        if years > 0 and first["net_assets"] > 0 and last["net_assets"] > 0:
            try:
                cagr = (math.pow(last["net_assets"] / first["net_assets"], 1 / years) - 1) * 100
            except (ValueError, ZeroDivisionError, OverflowError):
                pass

        # Family
        total = last["self_total"] + last["family_total"]
        family_ratio = last["family_total"] / total if total > 0 else 0
        spouse_ratio = last["spouse_total"] / total if total > 0 else 0

        delta_total = last["net_assets"] - first["net_assets"]
        delta_family = last["family_total"] - first["family_total"]
        family_growth_share = max(0, min(1, delta_family / delta_total)) if delta_total > 0 else 0

        # Gap
        gaps = conn.execute("""
            SELECT AVG(gap_pct) as avg_g, MAX(gap_pct) as max_g
            FROM real_estate_match
            WHERE politician_id = ? AND match_confidence >= 0.5
        """, (pid,)).fetchone()

        # Anomaly
        anomaly_score = 0
        flags = {}
        for y in yearly[1:]:
            if stdev_yoy > 0 and y["yoy_change"]:
                z = abs(y["yoy_change"] - median_yoy) / stdev_yoy
                if z > 2.5:
                    flags["sudden_spike"] = True
                    anomaly_score += min(20 + (z - 2.5) * 5, 30)
        anomaly_score = min(anomaly_score, 100)

        conn.execute("""
            INSERT OR REPLACE INTO derived_metrics
            (politician_id, analysis_start_year, analysis_end_year,
             cagr, absolute_growth, growth_rate,
             family_contribution_ratio, spouse_contribution_ratio, family_growth_share,
             avg_gap_pct, max_gap_pct,
             anomaly_flags, anomaly_score)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            pid, first["disclosure_year"], last["disclosure_year"],
            cagr, abs_growth, growth_rate,
            family_ratio, spouse_ratio, family_growth_share,
            gaps["avg_g"], gaps["max_g"],
            json.dumps(flags), anomaly_score,
        ))
        count += 1

    logger.info(f"파생 지표 {count}명 완료")
    return count


async def main():
    skip_api = "--skip-api" in sys.argv

    print("=" * 60)
    print("  실데이터 시딩: 2026 재산공개 + 실거래가 매칭")
    print(f"  엑셀: {EXCEL_PATH.name}")
    print(f"  실거래가 API: {'건너뜀' if skip_api else '연동'}")
    print("=" * 60)

    # DB 초기화 (기존 데이터 삭제)
    init_db()

    with get_db() as conn:
        # 기존 데이터 클리어
        for table in ["derived_metrics", "real_estate_match", "asset_disclosure_yearly", "asset_itemized", "politician_master"]:
            conn.execute(f"DELETE FROM {table}")

        # Step 1: 의원 + 자산
        name_to_id = seed_politicians(conn)
        seed_assets(conn, name_to_id)

        # Step 2: 합성 전년도
        create_synthetic_prior_year(conn)

        # Step 3: 연도별 집계
        compute_yearly_aggregates(conn)

        # Step 4: 실거래가 매칭
        if not skip_api:
            await match_real_estate(conn)

        # Step 5: 파생 지표
        compute_derived_metrics(conn)

    # 결과 확인
    with get_db() as conn:
        stats = {
            "의원 수": conn.execute("SELECT COUNT(*) as c FROM politician_master").fetchone()["c"],
            "자산 항목": conn.execute("SELECT COUNT(*) as c FROM asset_itemized").fetchone()["c"],
            "아파트 항목": conn.execute("SELECT COUNT(*) as c FROM asset_itemized WHERE kind LIKE '%아파트%'").fetchone()["c"],
            "연도별 집계": conn.execute("SELECT COUNT(*) as c FROM asset_disclosure_yearly").fetchone()["c"],
            "실거래 매칭": conn.execute("SELECT COUNT(*) as c FROM real_estate_match").fetchone()["c"],
            "파생 지표": conn.execute("SELECT COUNT(*) as c FROM derived_metrics").fetchone()["c"],
        }

    print("\n" + "=" * 60)
    print("  완료!")
    for k, v in stats.items():
        print(f"  {k}: {v}건")
    print("=" * 60)


if __name__ == "__main__":
    asyncio.run(main())
